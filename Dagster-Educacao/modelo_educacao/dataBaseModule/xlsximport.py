from openpyxl import load_workbook

class XlsxImport:
    @staticmethod
    def get_column_names(file_path):
        """
        Obtém os nomes das colunas do arquivo .xlsx.

        Args:
            file_path (str): Caminho para o arquivo .xlsx.

        Returns:
            list[str]: Lista de nomes das colunas.
        """
        workbook = load_workbook(file_path)
        sheet = workbook.active
        column_names = [cell.value for cell in sheet[1]]
        return column_names

    @staticmethod
    def get_data_rows(file_path):
        """
        Obtém as linhas de dados do arquivo .xlsx, excluindo o cabeçalho.

        Args:
            file_path (str): Caminho para o arquivo .xlsx.

        Returns:
            list[list[Any]]: Lista de linhas de dados.
        """
        workbook = load_workbook(file_path)
        sheet = workbook.active
        data_rows = [list(cell.value for cell in row) for row in sheet.iter_rows(min_row=2)]
        return data_rows

    @staticmethod
    def get_column_types(column_names):
        """
        Obtém os tipos de dados correspondentes das colunas do arquivo .xlsx.

        Args:
            column_names (list[str]): Lista de nomes das colunas.

        Returns:
            list[str]: Dicionário contendo o tipo de cada coluna.
        """
        column_types = []
        for column_name in column_names:
            column_type = input(f"Digite o tipo para a coluna '{column_name}': ")
            column_types.append(column_type)
        print(column_types)
        return column_types
